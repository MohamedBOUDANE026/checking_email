import smtplib
from email_validator import validate_email, EmailNotValidError  # type: ignore
import dns.resolver  # type: ignore
import sqlite3
import pandas as pd  # type: ignore
from openpyxl import load_workbook  # type: ignore

# Fonction pour créer la base de données et insérer des données aléatoires
def create_database():
    with sqlite3.connect('email_validation.db') as conn:
        c = conn.cursor()
        # Créer la table
        c.execute('''CREATE TABLE IF NOT EXISTS email_validation
                    (email TEXT, syntax_result TEXT, mx_result TEXT, deliverability_result TEXT)''')
        #print('Database created successfully')

# Étape 1 : Valider la syntaxe de l'email
def validate_email_syntax(email):
    try:
        valid = validate_email(email)
        return valid.email
    except EmailNotValidError as e:
        return f"Invalid email syntax: {e}"

#print('validate_email:')

# Étape 2 : Vérifier les enregistrements MX du domaine de l'email
def check_mx_record(domain):
    try:
        records = dns.resolver.resolve(domain, 'MX')
        mx_record = records[0].exchange
        return f"MX record found: {mx_record}"
    except (dns.resolver.NoAnswer, dns.resolver.NXDOMAIN, dns.resolver.Timeout) as e:
        return f"No MX record found for domain {domain}: {e}"
    except Exception as e:
        return f"An error occurred while checking MX record for domain {domain}: {e}"

#print('check mx record verified')

# Étape 3 : Vérifier la délivrabilité de l'email
def verify_email_deliverability(email):
    domain = email.split('@')[1]
    try:
        records = dns.resolver.resolve(domain, 'MX')
        mx_record = records[0].exchange.to_text()

        with smtplib.SMTP(mx_record, 25, timeout=30) as server:
            server.set_debuglevel(1)
            server.ehlo()
            server.starttls()
            server.ehlo()
            server.mail('test@example.com')
            code, message = server.rcpt(email)

            if code == 250:
                return f"Email {email} is deliverable."
            else:
                return f"Email {email} is not deliverable: {message}"
    except smtplib.SMTPConnectError as e:
        return f"SMTP connect error: {e}"
    except smtplib.SMTPServerDisconnected as e:
        return f"SMTP server disconnected: {e}"
    except smtplib.SMTPException as e:
        return f"SMTP error checking deliverability for {email}: {e}"
    except Exception as e:
        return f"Error checking deliverability for {email}: {e}"

#print('la délivrabilité d email verfied')

# Fonction pour enregistrer les résultats dans un fichier Excel
def store_results_in_csv(deliverable_results, non_deliverable_results):
    deliverable_df = pd.DataFrame(deliverable_results, columns=[
                                  'email', 'syntax_result', 'mx_result', 'deliverability_result'])
    non_deliverable_df = pd.DataFrame(non_deliverable_results, columns=[
                                      'email', 'syntax_result', 'mx_result', 'deliverability_result'])

    with pd.ExcelWriter('email_validation_results.xlsx', engine='openpyxl') as writer:
        deliverable_df.to_excel(writer, sheet_name='deliverable', index=False)
        non_deliverable_df.to_excel(
            writer, sheet_name="Non deliverable", index=False)

    # Charger le fichier Excel pour ajuster la largeur des colonnes
    wb = load_workbook('email_validation_results.xlsx')

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for column_cells in sheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = length + 2

    wb.save('email_validation_results.xlsx')

#print('store results in csv verified')

# Fonction principale pour valider un email et stocker les résultats
def validate_email_address(email, deliverable_results, non_deliverable_results):
    syntax_result = validate_email_syntax(email)
    #print(f"Syntax validation result: {syntax_result}")

    if "Invalid email syntax" in syntax_result:
        mx_result = None
        deliverability_result = None
        non_deliverable_results.append(
            [email, syntax_result, mx_result, deliverability_result])
    else:
        domain = email.split('@')[1]
        mx_result = check_mx_record(domain)
        #print(f"MX record check result: {mx_result}")

        if "No MX record found" in mx_result:
            deliverability_result = None
            non_deliverable_results.append(
                [email, syntax_result, mx_result, deliverability_result])
        else:
            deliverability_result = verify_email_deliverability(email)
            #print(f"Deliverability check result: {deliverability_result}")
            if "is deliverable" in deliverability_result:
                deliverable_results.append(
                    [email, syntax_result, mx_result, deliverability_result])
            else:
                non_deliverable_results.append(
                    [email, syntax_result, mx_result, deliverability_result])

# Créer la base de données avec des données aléatoires
create_database()

# Tester les fonctions avec les emails de la base de données
if __name__ == "__main__":
    deliverable_results = []
    non_deliverable_results = []

    with sqlite3.connect('email_validation.db') as conn:
        c = conn.cursor()
        c.execute("SELECT email FROM email_validation")
        emails = c.fetchall()

        for email in emails:
            validate_email_address(
                email[0], deliverable_results, non_deliverable_results)

    # Afficher les résultats dans le terminal
    print("\nDeliverable Emails:")
    for result in deliverable_results:
        print(result[0])

    print("\nNon-Deliverable Emails:")
    for result in non_deliverable_results:
        print(result[0])

    # Stocker les résultats dans un fichier Excel
    store_results_in_csv(deliverable_results, non_deliverable_results)
